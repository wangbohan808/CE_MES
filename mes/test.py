import asyncio
import edge_tts
import os
from openpyxl import load_workbook

# fr-FR: 法语(法国)
# de-DE: 德语(德国)
# es-ES: 西班牙语(西班牙)
# ja-JP: 日语(日本)
# en-GB: 英语(英国)
# en-US: 英语(美国)
# zh-HK: 中文(粤语，繁体)
# zh-TW: 中文(台湾普通话)
# zh-CN: 中文(普通话，简体)
# it-IT: 意大利语（意大利）

set_lange = 'it-IT'
set_debug = 0


async def get_voices(locale=""):
    # 获取所有语音列表
    all_voices = await edge_tts.list_voices()
    my_need_voices = []
    for voice in all_voices:
        try:
            get_locale = voice.get("Locale")
            # 筛选中文语音
            if get_locale and get_locale.startswith(locale):
                my_need_voices.append({
                    "locale": locale,
                    "description": voice.get("ShortName"),
                    "Gender": voice.get("Gender")
                })
                print({
                    "locale": locale,
                    "description": voice.get("ShortName"),
                    "Gender": voice.get("Gender")
                })
        except Exception as e:
            print(f"处理语音时出错: {e}")
            continue

    return my_need_voices


async def text_to_speech(text, voice_name="zh-CN-YunxiNeural", output_file="output.mp3"):
    """将文本转换为语音并保存为文件

    Args:
        text: 要转换的文本内容
        voice_name: 语音名称，默认为中文女性语音
        output_file: 输出文件名，默认为 "output.mp3"
    """
    # 创建 TTS 通信对象
    communicate = edge_tts.Communicate(text, voice_name)

    # 保存为音频文件
    print(f"正在生成语音: {voice_name}")
    await communicate.save(output_file)
    print(f"语音已保存至: {os.path.abspath(output_file)}")


def read_voices_text(lg=""):
    wb = load_workbook('乐动公版音效多国语音0302.xlsx')
    ws = wb.active

    # 获取列名并找到目标列
    column_names = [cell.value for cell in ws[1]]
    target_col = None

    # 查找法语列（第一行遍历）
    for idx, lg_name in enumerate(column_names, 1):  # 从1开始计数
        if lg_name and lg_name.startswith(lg):
            target_col = idx
            break

    if not target_col:
        raise ValueError(f"工作表中未找到 {lg} 列")

    # 构建字典（第一列 -> 法语列）
    data_dict = {}
    for row in ws.iter_rows(min_row=2):  # 从第二行开始遍历
        key_cell = row[0]  # 第一列（A列）
        value_cell = row[target_col - 1]  # 法语列（列索引转换从0开始）

        # 处理空值情况（可选）
        key = key_cell.value if key_cell.value else f"空键_{row[0].row}"
        value = value_cell.value if value_cell.value else None
        if value and key:
            # 处理重复键（后出现的覆盖前者）
            data_dict[key] = value

    # 验证输出
    print("生成的字典示例：")
    for k, v in list(data_dict.items())[:5]:  # 打印前3项
        print(f"{k}: {v}")
    return data_dict


# 运行异步函数
if __name__ == "__main__":
    voices = asyncio.run(get_voices(set_lange))
    text_dirt = read_voices_text("意大利")

    for key, value in text_dirt.items():
        name = key
        text = value
        if set_debug:
            # 测试
            asyncio.run(text_to_speech(text, "it-IT-ElsaNeural", "Q003-1.mp3"))
            break
        else:
            # 正常生成  fr-FR-DeniseNeural   de-DE-AmalaNeural es-ES-ElviraNeural
            asyncio.run(text_to_speech(text, "it-IT-ElsaNeural", "out-it/" + name + ".mp3"))

