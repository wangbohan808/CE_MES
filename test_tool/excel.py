from openpyxl import load_workbook
from openpyxl import Workbook
import csv


"""
向Excel文件末尾添加一条记录
参数:
    file_path: Excel文件路径
    title: 标题列表（如果文件不存在则创建） ["姓名", "年龄"]
    data: 要添加的数据列表（与标题对应）    ["王五", “27”]
"""


def add_record_to_excel(file_path, title, data):

    try:
        # 尝试加载现有工作簿
        wb = load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        # 如果文件不存在，创建新工作簿
        wb = Workbook()
        sheet = wb.active
        sheet.append(title)  # 添加标题行

    sheet.append(data)  # 在末尾添加数据
    try:
        wb.save(file_path)
        return True, "保存成功"
    except PermissionError:
        return False, "文件被占用或无写入权限"
    except Exception as e:
        return False, f"未知错误: {str(e)}"


def add_record_to_csv(file_path, record):

    try:
        with open(file_path, 'a', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=record.keys())

            if f.tell() == 0:
                writer.writeheader()
            writer.writerow(record)
    except PermissionError:
        print("文件被占用，写入csv文件失败")
    except Exception as e:
        print(f"未知错误：{str(e)}")



















